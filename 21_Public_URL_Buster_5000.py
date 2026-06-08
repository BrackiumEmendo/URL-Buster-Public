import asyncio
import datetime
import random
import time
import re
from pathlib import Path
from typing import List, Optional
from collections import defaultdict
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

import pandas as pd
import streamlit as st
import requests
import snowflake.connector
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from playwright.async_api import async_playwright

# --------------------------------------------------------------------------- #
# CONFIG
# --------------------------------------------------------------------------- #
AUDIT_DIR = Path("audit")
AUDIT_DIR.mkdir(exist_ok=True)

CHECKPOINT_FILE  = AUDIT_DIR / "checkpoint.csv"
URLLIST_FILE     = AUDIT_DIR / "urllist.csv"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/122.0 Safari/537.36"
)

MAX_CONCURRENT   = 2
HTTP_TIMEOUT     = 10
RETRIES          = 2
DOMAIN_DELAY     = 1.5
BATCH_SIZE       = 5000
CHECKPOINT_EVERY = 10

TRACKING_PARAMS = {
    'cid', 'src', 'its', 'l', 'f', 'step', 'site', 'fh', 'sign', 'add',
    'ppid', 'platform', 'see', 'locale', 'lang', 'type', 'sub', 'filter',
    'chang', 'i', 'daddr', 'target', 'show', 'area', 'caller', 'product',
    'device-type',
}

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT})

domain_last_hit = defaultdict(float)


# --------------------------------------------------------------------------- #
# SNOWFLAKE
# --------------------------------------------------------------------------- #
def get_snowflake_connection(account, user, warehouse, database, schema):
    return snowflake.connector.connect(
        account=account,
        user=user,
        authenticator="externalbrowser",
        warehouse=warehouse,
        database=database,
        schema=schema,
    )


def pull_urls_from_snowflake(conn, query: str) -> List[str]:
    df = pd.read_sql(query, conn)
    return df.iloc[:, 0].dropna().tolist()


def check_scraped_in_snowflake(conn, urls: List[str], concepts_table: str) -> pd.DataFrame:
    sql = build_snowflake_check_query(urls, concepts_table)
    df  = pd.read_sql(sql, conn)
    df.columns = [c.lower() for c in df.columns]
    return df


# --------------------------------------------------------------------------- #
# CHECKPOINT
# --------------------------------------------------------------------------- #
def save_checkpoint(results: list, batch_idx: int):
    path = AUDIT_DIR / f"checkpoint_batch{batch_idx}.csv"
    pd.DataFrame(results).to_csv(path, index=False)


def load_checkpoint(batch_idx: int) -> list:
    path = AUDIT_DIR / f"checkpoint_batch{batch_idx}.csv"
    if path.exists():
        return pd.read_csv(path).fillna("").to_dict("records")
    return []


def clear_checkpoint(batch_idx: int):
    path = AUDIT_DIR / f"checkpoint_batch{batch_idx}.csv"
    if path.exists():
        path.unlink()


def clear_all_checkpoints():
    for f in AUDIT_DIR.glob("checkpoint_batch*.csv"):
        f.unlink()
    if URLLIST_FILE.exists():
        URLLIST_FILE.unlink()


def save_urllist(urls: List[str]):
    pd.DataFrame({"url": urls}).to_csv(URLLIST_FILE, index=False)


def load_urllist() -> List[str]:
    if URLLIST_FILE.exists():
        return pd.read_csv(URLLIST_FILE)["url"].dropna().tolist()
    return []


def any_checkpoint_exists() -> bool:
    return any(AUDIT_DIR.glob("checkpoint_batch*.csv"))


def get_saved_batch_indices() -> List[int]:
    indices = []
    for f in AUDIT_DIR.glob("checkpoint_batch*.csv"):
        try:
            idx = int(f.stem.replace("checkpoint_batch", ""))
            indices.append(idx)
        except ValueError:
            pass
    return sorted(indices)


# --------------------------------------------------------------------------- #
# URL CLEANER
# --------------------------------------------------------------------------- #
def clean_apple_url(url: str) -> str:
    url = url.strip()
    if not url:
        return url
    return re.sub(
        r"/\d+(?:\.\d+)*/([a-z][a-z0-9-]*)(?:/\d+(?:\.\d+)*)?",
        r"/\1",
        url,
        flags=re.IGNORECASE,
    )


def strip_tracking_params(url: str) -> str:
    try:
        parsed = urlparse(url)
        if not parsed.query:
            return url
        params    = parse_qs(parsed.query, keep_blank_values=True)
        cleaned   = {k: v for k, v in params.items()
                     if k.lower() not in TRACKING_PARAMS}
        new_query = urlencode(cleaned, doseq=True)
        return urlunparse(parsed._replace(query=new_query))
    except Exception:
        return url


def get_platform_url(original: str, redirect: str) -> str:
    original = original.strip() if original else ""
    cleaned_original = clean_apple_url(original)
    if cleaned_original != original:
        return cleaned_original
    return ""


def get_platform_redirect_url(platform_url: str, original: str, redirect: str) -> str:
    platform_url = platform_url.strip() if platform_url else ""
    original     = original.strip() if original else ""
    redirect     = redirect.strip() if redirect else ""
    if not platform_url:
        return ""
    if not redirect:
        return ""
    if redirect != platform_url:
        return redirect
    return ""


def get_final_url(platform_redirect_url: str, platform_url: str,
                  redirect_url: str, original_url: str) -> str:
    for candidate in [platform_redirect_url, platform_url, redirect_url, original_url]:
        val = str(candidate).strip() if candidate else ""
        if val:
            return strip_tracking_params(clean_apple_url(val))
    return ""


def get_snowflake_url(platform_url: str, redirect_url: str, original_url: str) -> str:
    if platform_url and isinstance(platform_url, str) and platform_url.strip():
        return platform_url.strip()
    source = redirect_url.strip() if redirect_url and redirect_url.strip() else original_url.strip()
    return clean_apple_url(source)


# --------------------------------------------------------------------------- #
# SNOWFLAKE QUERY BUILDER
# --------------------------------------------------------------------------- #
def build_snowflake_check_query(urls: List[str], concepts_table: str) -> str:
    if not urls or not concepts_table:
        return ""

    select_blocks = []
    for i, u in enumerate(urls):
        u_safe = u.replace("'", "''")
        if i == 0:
            block = (
                "    SELECT\n"
                "        ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS rn,\n"
                f"        '{u_safe}' AS url"
            )
        else:
            block = (
                "    SELECT\n"
                "        ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS rn,\n"
                f"        '{u_safe}'"
            )
        select_blocks.append(block)

    cte_body = "\n    UNION ALL\n".join(select_blocks)
    return (
        "WITH url_list AS (\n"
        f"{cte_body}\n"
        ")\n"
        "SELECT\n"
        "    u.url,\n"
        "    CASE WHEN COUNT(t.source) > 0 THEN 1 ELSE 0 END AS in_table,\n"
        "    COUNT(t.source) AS match_count,\n"
        "    MAX(t.DEPRECATED) AS deprecated\n"
        "FROM url_list u\n"
        f"LEFT JOIN {concepts_table} t\n"
        "       ON u.url = t.source\n"
        "GROUP BY u.rn, u.url\n"
        "ORDER BY u.rn;"
    )


# --------------------------------------------------------------------------- #
# DEDUPLICATION
# --------------------------------------------------------------------------- #
def deduplicate_needs_scraping(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df["Scrape URL"] = df["Scrape URL"].astype(str).str.strip()
    df = df[df["Scrape URL"] != ""].copy()
    df = df[df["Scrape URL"].str.lower() != "nan"].copy()
    df = df.drop_duplicates(subset=["Scrape URL"], keep="first").reset_index(drop=True)

    seen_titles: set = set()
    keep: List[bool] = []

    for _, row in df.iterrows():
        title     = str(row.get("Title", "")).strip()
        title_key = title.lower()

        if title and title_key not in ("", "nan") and title in seen_titles:
            keep.append(False)
            continue

        if title and title_key not in ("", "nan"):
            seen_titles.add(title)
        keep.append(True)

    return df[keep].reset_index(drop=True)


# --------------------------------------------------------------------------- #
# EXCEL EXPORT
# --------------------------------------------------------------------------- #
def write_excel(df_all: pd.DataFrame, path: Path):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    blue_fill   = PatternFill("solid", fgColor="BDD7EE")

    ws1 = wb.active
    ws1.title = "All URLs"
    for r in dataframe_to_rows(df_all, index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font

    in_table_post_idx = next(
        (cell.column for cell in ws1[1] if cell.value == "In Table Post"), None
    )
    in_table_pre_idx  = next(
        (cell.column for cell in ws1[1] if cell.value == "In Table Pre"), None
    )
    url_col_idx       = next(
        (cell.column for cell in ws1[1] if cell.value == "URL"), None
    )
    final_url_col_idx = next(
        (cell.column for cell in ws1[1] if cell.value == "Final URL"), None
    )

    for row in ws1.iter_rows(min_row=2):
        in_table_post_val = (
            str(row[in_table_post_idx - 1].value).strip()
            if in_table_post_idx else ""
        )

        if in_table_post_val == "0":
            no_redirect = False
            if url_col_idx and final_url_col_idx:
                orig_val    = str(row[url_col_idx - 1].value).strip()
                final_val   = str(row[final_url_col_idx - 1].value).strip()
                no_redirect = (orig_val == final_val)

            fill = red_fill if no_redirect else yellow_fill
            for cell in row:
                cell.fill = fill

        if in_table_pre_idx and url_col_idx:
            in_table_pre_val = str(row[in_table_pre_idx - 1].value).strip()
            if in_table_pre_val == "0":
                row[url_col_idx - 1].fill = blue_fill

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    in_table_col = "In Table Post" if "In Table Post" in df_all.columns else None
    if in_table_col:
        mask = (
            (df_all[in_table_col].astype(str) == "0") &
            (df_all["Error"].astype(str).str.strip() == "")
        )
        needs = df_all[mask].copy()
    else:
        needs = df_all[df_all["Error"].astype(str).str.strip() == ""].copy()

    has_final = needs["Final URL"].astype(str).str.strip() != ""
    needs = needs[has_final].copy()
    needs["Scrape URL"] = needs["Final URL"]

    needs_out = needs[["Scrape URL", "Title"]].reset_index(drop=True)
    needs_out = deduplicate_needs_scraping(needs_out)

    ws2 = wb.create_sheet(title="Needs Scraping")
    for r in dataframe_to_rows(needs_out, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    wb.save(path)


# --------------------------------------------------------------------------- #
# HELPERS
# --------------------------------------------------------------------------- #
def now():
    return datetime.datetime.now().strftime("%Y%m%d_%H%M")


async def wait_for_domain(url: str):
    domain   = url.split("/")[2]
    now_time = time.time()
    elapsed  = now_time - domain_last_hit[domain]
    if elapsed < DOMAIN_DELAY:
        await asyncio.sleep(DOMAIN_DELAY - elapsed)
    domain_last_hit[domain] = time.time()


# --------------------------------------------------------------------------- #
# CONTENT DETECTION
# --------------------------------------------------------------------------- #
def has_real_guide_content(html: str) -> bool:
    if not html:
        return False
    lower   = html.lower()
    signals = [
        "<h1", "role=\"main\"", "table of contents", "learn more",
        "mainstage", "logic pro", "mac help", "controls",
        "parameters", "midi", "audio",
    ]
    return sum(1 for s in signals if s in lower) >= 3


def has_not_found_text(html: str) -> bool:
    if not html:
        return False
    lower   = html.lower()
    signals = [
        "the page you're looking for can't be found",
        "we can't find the page",
        "sosumi-not-found",
    ]
    return any(s in lower for s in signals)


def is_probably_throttled(html: str, status: int) -> bool:
    if not html:
        return True
    if status == 503:
        return True
    if status == 200 and len(html) < 3000:
        return True
    lower = html.lower()
    if "akamai" in lower or "access denied" in lower:
        return True
    return False


def is_real_404(html: str) -> bool:
    if not html:
        return False
    lower = html.lower()
    if not has_not_found_text(html):
        return False
    if "table of contents" in lower:
        return False
    if "role=\"main\"" in lower:
        return False
    if "<h1" in lower and "not found" not in lower:
        return False
    return True


# --------------------------------------------------------------------------- #
# TITLE EXTRACTION
# --------------------------------------------------------------------------- #
def extract_title(html: str) -> str:
    if not html:
        return ""
    m     = re.search(r"<title[^>]*>(.*?)</title>", html, re.I | re.S)
    title = m.group(1).strip() if m else ""
    if not title or "page not found" in title.lower():
        og = re.search(r'<meta property="og:title" content="(.*?)"', html, re.I)
        if og:
            title = og.group(1).strip()
    if not title or "page not found" in title.lower():
        h1 = re.search(r"<h1[^>]*>(.*?)</h1>", html, re.I | re.S)
        if h1:
            clean = re.sub("<.*?>", "", h1.group(1)).strip()
            if clean:
                title = clean
    return title


def fix_title(url: str, title: str, html: str) -> str:
    if "support.apple.com/guide/" in url:
        if has_real_guide_content(html) and "page not found" in title.lower():
            h1 = re.search(r"<h1[^>]*>(.*?)</h1>", html, re.I | re.S)
            if h1:
                clean = re.sub("<.*?>", "", h1.group(1)).strip()
                if clean:
                    return clean
        if not has_real_guide_content(html) and has_not_found_text(html):
            return "Page Not Found"
    return title


# --------------------------------------------------------------------------- #
# CLASSIFIER
# --------------------------------------------------------------------------- #
def classify(url: str, html: str, status: int) -> Optional[str]:
    if is_probably_throttled(html, status):
        return None
    if has_real_guide_content(html):
        return None
    if is_real_404(html) and status == 404:
        return "404"
    if status == 404 and not has_not_found_text(html):
        return None
    if status == 404:
        return "404"
    if status >= 400:
        return str(status)
    return None


# --------------------------------------------------------------------------- #
# FETCH
# --------------------------------------------------------------------------- #
def fast_fetch(url: str):
    for attempt in range(RETRIES + 1):
        try:
            r        = session.get(url, timeout=HTTP_TIMEOUT, allow_redirects=True)
            html     = r.text
            title    = extract_title(html)
            title    = fix_title(url, title, html)
            redirect = r.url if r.url != url else ""
            if redirect:
                redirect = strip_tracking_params(redirect)
            return {"url": url, "redirect_url": redirect,
                    "status": r.status_code, "html": html, "title": title}
        except Exception:
            if attempt == RETRIES:
                return {"url": url, "redirect_url": "", "status": "ERROR",
                        "html": "", "title": "", "error": "ERROR"}
            time.sleep(random.uniform(1.0, 2.5) * (attempt + 1))


def needs_browser(res) -> bool:
    html   = res.get("html", "")
    status = res.get("status", 0)
    if status in [401, 403, 429]:
        return True
    if is_probably_throttled(html, status):
        return True
    if status == 200 and len(html) < 1500:
        return True
    return False


# --------------------------------------------------------------------------- #
# PLAYWRIGHT
# --------------------------------------------------------------------------- #
async def browser_fetch(url: str, context):
    page = await context.new_page()
    try:
        response  = await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(random.uniform(0.5, 1.2))
        html      = await page.content()
        title     = await page.title()
        title     = fix_title(url, title, html)
        final_url = page.url
        final_url = strip_tracking_params(final_url)
        status    = response.status if response else 0
        error     = classify(url, html, status)
        redirect  = final_url if final_url != url else ""
        return {"url": url, "redirect_url": redirect,
                "title": title, "error": error or ""}
    except Exception:
        return {"url": url, "redirect_url": "", "title": "", "error": "ERROR"}
    finally:
        await page.close()


# --------------------------------------------------------------------------- #
# PLATFORM URL CHECKER
# --------------------------------------------------------------------------- #
async def check_platform_urls(platform_urls: List[str], context):
    unique    = [u for u in dict.fromkeys(platform_urls) if u]
    results   = {}
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)

    async def worker(url):
        async with semaphore:
            await wait_for_domain(url)
            await asyncio.sleep(random.uniform(0.8, 1.8))

            fast = await asyncio.to_thread(fast_fetch, url)

            if needs_browser(fast) or not fast.get("redirect_url"):
                res = await browser_fetch(url, context)
            else:
                res = {
                    "url":          url,
                    "redirect_url": fast.get("redirect_url", ""),
                    "title":        fast.get("title", ""),
                    "error":        classify(
                                        url,
                                        fast.get("html", ""),
                                        fast.get("status", 0),
                                    ) or "",
                }

            results[url] = res

    await asyncio.gather(*[worker(u) for u in unique])
    return results


# --------------------------------------------------------------------------- #
# CRAWLER
# --------------------------------------------------------------------------- #
async def crawl(
    urls: List[str],
    batch_idx: int,
    existing_results: list = None,
    platform_check: bool = True,
):
    done_urls = set()
    results   = []

    if existing_results:
        results   = existing_results
        done_urls = {r["url"] for r in results}

    remaining = [u for u in urls if u not in done_urls]

    if not remaining:
        return results, {}

    semaphore  = asyncio.Semaphore(MAX_CONCURRENT)
    start_time = time.time()
    completed  = len(results)
    total      = len(urls)

    progress_bar = st.session_state.get("progress")
    status_text  = st.session_state.get("status_text")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(user_agent=USER_AGENT)

        async def worker(url):
            async with semaphore:
                await wait_for_domain(url)
                await asyncio.sleep(random.uniform(1.0, 2.5))

                fast = await asyncio.to_thread(fast_fetch, url)

                if needs_browser(fast):
                    return await browser_fetch(url, context)

                error = classify(url, fast.get("html", ""), fast.get("status", 0))

                if error in ["404", "503"]:
                    retry       = await asyncio.to_thread(fast_fetch, url)
                    retry_error = classify(url, retry.get("html", ""), retry.get("status", 0))
                    if retry_error is None:
                        return {"url": url,
                                "redirect_url": retry.get("redirect_url", ""),
                                "title":        retry.get("title", ""),
                                "error":        ""}
                    return await browser_fetch(url, context)

                if error is None:
                    return {"url": url,
                            "redirect_url": fast.get("redirect_url", ""),
                            "title":        fast.get("title", ""),
                            "error":        ""}

                return await browser_fetch(url, context)

        try:
            tasks = [worker(u) for u in remaining]
            for i, coro in enumerate(asyncio.as_completed(tasks), 1):
                res = await coro
                results.append(res)
                completed += 1

                if i % CHECKPOINT_EVERY == 0:
                    save_checkpoint(results, batch_idx)

                elapsed = time.time() - start_time
                rate    = i / elapsed if elapsed else 0
                eta     = (len(remaining) - i) / rate if rate else 0

                if progress_bar:
                    progress_bar.progress(
                        completed / (total * 2) if platform_check else completed / total
                    )
                if status_text:
                    status_text.text(
                        f"Pass 1 — Batch {batch_idx} — {completed:,}/{total:,} | "
                        f"{rate:.1f} URLs/s | ETA {int(eta)}s"
                    )

        except (KeyboardInterrupt, asyncio.CancelledError, Exception) as exc:
            save_checkpoint(results, batch_idx)
            if status_text:
                status_text.warning(
                    f"Batch {batch_idx} interrupted at {completed:,} URLs — "
                    f"checkpoint saved."
                )
            raise exc

        finally:
            save_checkpoint(results, batch_idx)

        platform_check_results = {}

        if platform_check:
            if status_text:
                status_text.text(
                    f"Pass 2 — Batch {batch_idx} — Checking Platform URLs…"
                )

            pass1_map = {r["url"]: r for r in results}

            platform_urls_to_check = []
            for url in urls:
                r            = pass1_map.get(url, {})
                raw_redirect = r.get("redirect_url", "")
                purl         = get_platform_url(url, raw_redirect)
                if purl and purl != url:
                    platform_urls_to_check.append(purl)

            platform_check_results = await check_platform_urls(
                platform_urls_to_check, context
            )

        if progress_bar:
            progress_bar.progress(1.0)
        if status_text:
            if platform_check:
                status_text.text(
                    f"Done — Batch {batch_idx} — {total:,} original + "
                    f"{len(platform_check_results):,} platform URLs checked"
                )
            else:
                status_text.text(
                    f"Done — Batch {batch_idx} — {total:,} URLs checked "
                    f"(simple mode — no Platform URL pass)"
                )

        await browser.close()

    return results, platform_check_results


# --------------------------------------------------------------------------- #
# PROCESS ONE BATCH
# --------------------------------------------------------------------------- #
def process_batch(
    batch_urls: List[str],
    batch_idx: int,
    num_batches: int,
    timestamp: str,
    resume: bool,
    platform_check: bool = True,
    download_placeholder=None,
    concepts_table: str = "",
) -> Optional[Path]:

    st.markdown(f"---\n### Batch {batch_idx} of {num_batches} — {len(batch_urls):,} URLs")

    if not platform_check:
        st.caption("Simple mode — Platform URL cleanup and Pass 2 check skipped.")

    existing = load_checkpoint(batch_idx) if resume else []
    if existing:
        st.info(f"Resuming batch {batch_idx} from {len(existing):,} already crawled.")

    st.session_state.progress    = st.progress(0)
    st.session_state.status_text = st.empty()

    try:
        results, platform_check_results = asyncio.run(
            crawl(
                batch_urls,
                batch_idx,
                existing_results=existing,
                platform_check=platform_check,
            )
        )
    except Exception as e:
        st.error(
            f"Batch {batch_idx} stopped: `{e}`\n\n"
            "Progress saved — use **Resume** to continue."
        )
        return None

    clear_checkpoint(batch_idx)

    out = pd.DataFrame(results)
    for col in ["url", "redirect_url", "title", "error"]:
        if col not in out.columns:
            out[col] = ""
    out = out.fillna("")

    out["raw_redirect"] = out["redirect_url"]

    if platform_check:
        out["platform_url"] = out.apply(
            lambda row: get_platform_url(row["url"], row["raw_redirect"]), axis=1
        )

        def resolve_platform_redirect(row):
            purl = row["platform_url"]
            if not purl:
                return ""
            check = platform_check_results.get(purl)
            if not check:
                return ""
            actual_redirect = check.get("redirect_url", "")
            if actual_redirect and actual_redirect != purl:
                return actual_redirect
            return ""

        out["platform_redirect_url"] = out.apply(resolve_platform_redirect, axis=1)

        def resolve_title(row):
            purl = row["platform_url"]
            if purl and purl in platform_check_results:
                t = platform_check_results[purl].get("title", "")
                if t:
                    return t
            return row["title"]

        out["title"] = out.apply(resolve_title, axis=1)

        out["redirect_url"] = out.apply(
            lambda row: row["raw_redirect"]
            if (row["raw_redirect"] and not row["platform_url"])
            else "",
            axis=1,
        )
        out = out.drop(columns=["raw_redirect"])

        out["final_url"] = out.apply(
            lambda row: get_final_url(
                row["platform_redirect_url"],
                row["platform_url"],
                row["redirect_url"],
                row["url"],
            ),
            axis=1,
        )

    else:
        out["platform_url"]          = ""
        out["platform_redirect_url"] = ""
        out["redirect_url"]          = out["raw_redirect"]
        out = out.drop(columns=["raw_redirect"])

        out["final_url"] = out.apply(
            lambda row: row["redirect_url"] if row["redirect_url"] else row["url"],
            axis=1,
        )

    out = out.rename(columns={
        "url":                   "URL",
        "platform_url":          "Platform URL",
        "platform_redirect_url": "Platform Redirect URL",
        "redirect_url":          "Redirect URL",
        "final_url":             "Final URL",
        "title":                 "Title",
        "error":                 "Error",
    })

    out = out[[
        "URL", "Platform URL", "Platform Redirect URL",
        "Redirect URL", "Final URL", "Title", "Error"
    ]]

    original_order = pd.DataFrame({"URL": batch_urls})
    out = original_order.merge(out, on="URL", how="left").fillna("")

    out["Final URL"] = out["Final URL"].apply(
        lambda u: strip_tracking_params(str(u).strip()) if str(u).strip() else u
    )

    original_urls  = out["URL"].tolist()
    snowflake_urls = out["Final URL"].tolist()

    if "sf_conn" in st.session_state and concepts_table:
        try:
            with st.spinner(f"Snowflake scrape-check for batch {batch_idx}..."):
                scraped_post = check_scraped_in_snowflake(
                    st.session_state.sf_conn, snowflake_urls, concepts_table
                )
                scraped_pre = check_scraped_in_snowflake(
                    st.session_state.sf_conn, original_urls, concepts_table
                )

            scraped_post.columns = [c.lower() for c in scraped_post.columns]
            scraped_post = scraped_post.rename(columns={
                "url":         "Snowflake URL",
                "in_table":    "In Table Post",
                "match_count": "Match Count Post",
                "deprecated":  "Deprecated",
            })
            scraped_post["Snowflake URL"] = (
                scraped_post["Snowflake URL"].astype(str).str.strip()
            )
            sf_post        = scraped_post.set_index("Snowflake URL")[
                ["In Table Post", "Match Count Post", "Deprecated"]
            ]
            out["_sf_post"] = snowflake_urls
            out = out.merge(sf_post, left_on="_sf_post", right_index=True, how="left")
            out = out.drop(columns=["_sf_post"])

            scraped_pre.columns = [c.lower() for c in scraped_pre.columns]
            scraped_pre = scraped_pre.rename(columns={
                "url":         "Snowflake URL Pre",
                "in_table":    "In Table Pre",
                "match_count": "Match Count Pre",
                "deprecated":  "_dep_pre",
            })
            scraped_pre["Snowflake URL Pre"] = (
                scraped_pre["Snowflake URL Pre"].astype(str).str.strip()
            )
            sf_pre         = scraped_pre.set_index("Snowflake URL Pre")[
                ["In Table Pre", "Match Count Pre"]
            ]
            out["_sf_pre"] = original_urls
            out = out.merge(sf_pre, left_on="_sf_pre", right_index=True, how="left")
            out = out.drop(columns=["_sf_pre"])

            for col in ["In Table Post", "Match Count Post", "Deprecated",
                        "In Table Pre",  "Match Count Pre"]:
                out[col] = out[col].fillna("").astype(str)

            st.success(f"Batch {batch_idx} scrape-check complete.")
        except Exception as e:
            st.warning(f"Batch {batch_idx} scrape-check skipped — {e}")
    else:
        if "sf_conn" not in st.session_state:
            st.warning(
                f"Batch {batch_idx} — Snowflake not connected, scrape-check skipped."
            )
        elif not concepts_table:
            st.warning(
                f"Batch {batch_idx} — No concepts table provided, scrape-check skipped. "
                "Enter your table name in the sidebar."
            )

    path = (
        AUDIT_DIR / f"run-{timestamp}.xlsx"
        if num_batches == 1
        else AUDIT_DIR / f"run-{timestamp}-batch{batch_idx}of{num_batches}.xlsx"
    )
    write_excel(out, path)

    csv_path = path.with_suffix(".csv")
    out.to_csv(csv_path, index=False)

    st.success(f"Batch {batch_idx} complete — `{path.name}`")

    if download_placeholder is not None:
        with open(path, "rb") as f:
            download_placeholder.download_button(
                label=f"Download Batch {batch_idx} — {path.name}",
                data=f,
                file_name=path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_batch_{batch_idx}_{timestamp}",
            )

    return path


# --------------------------------------------------------------------------- #
# UI
# --------------------------------------------------------------------------- #
st.set_page_config(layout="wide")
st.title("URL Buster 5000")

# ── Snowflake sidebar ────────────────────────────────────────────────────────
with st.sidebar:
    st.header("Snowflake")
    sf_account        = st.text_input("Account",         value="")
    sf_user           = st.text_input("User",            value="")
    sf_warehouse      = st.text_input("Warehouse",       value="")
    sf_database       = st.text_input("Database",        value="")
    sf_schema         = st.text_input("Schema",          value="")
    sf_concepts_table = st.text_input(
        "Concepts Table (fully qualified)",
        value="",
        placeholder="YOUR_DB.YOUR_SCHEMA.YOUR_TABLE",
        help="The fully qualified table name used for the scrape-check, e.g. MY_DB.MY_SCHEMA.MY_CONCEPTS_TABLE"
    )

    st.caption("Make sure you are on **DCVPN** before connecting.")
    connect_btn = st.button("Connect (SSO pop-up)", use_container_width=True)

    if connect_btn:
        if not all([sf_account, sf_user, sf_warehouse, sf_database, sf_schema]):
            st.error("Fill in all connection fields first.")
        else:
            try:
                with st.spinner("Opening SSO browser window..."):
                    conn = get_snowflake_connection(
                        sf_account, sf_user, sf_warehouse, sf_database, sf_schema
                    )
                    st.session_state.sf_conn           = conn
                    st.session_state.sf_concepts_table = sf_concepts_table
                st.success("Connected!")
            except Exception as e:
                st.error(f"Connection failed: {e}")

    st.caption(
        "Snowflake connected" if "sf_conn" in st.session_state
        else "Not connected"
    )

# ── Tabs ─────────────────────────────────────────────────────────────────────
tab_pull, tab_csv = st.tabs(["Pull from Snowflake", "Upload CSV instead"])

with tab_pull:
    st.subheader("Pull Query")
    st.caption("Paste your Snowflake query below. The first column returned will be used as the URL list.")
    pull_query = st.text_area(
        "Snowflake Query",
        value="",
        height=300,
        placeholder="SELECT url FROM your_database.your_schema.your_table WHERE ..."
    )
    pull_btn = st.button("Pull URLs from Snowflake", use_container_width=True)

    if pull_btn:
        if "sf_conn" not in st.session_state:
            st.error("Connect to Snowflake first using the sidebar.")
        elif not pull_query.strip():
            st.error("Please enter a query before pulling.")
        else:
            try:
                with st.spinner("Running pull query..."):
                    pulled = pull_urls_from_snowflake(st.session_state.sf_conn, pull_query)
                    st.session_state.urls = pulled
                    save_urllist(pulled)
                st.success(f"Pulled {len(pulled):,} URLs")
                st.dataframe(pd.DataFrame({"URL": pulled[:20]}), use_container_width=True)
                if len(pulled) > 20:
                    st.caption(f"Showing first 20 of {len(pulled):,}")
            except Exception as e:
                st.error(f"Pull failed: {e}")

with tab_csv:
    uploaded = st.file_uploader("Upload CSV")
    if uploaded:
        df_up = pd.read_csv(uploaded)
        st.session_state.urls = df_up.iloc[:, 0].dropna().tolist()
        save_urllist(st.session_state.urls)
        st.success(f"Loaded {len(st.session_state.urls):,} URLs from CSV")
        if "sf_conn" not in st.session_state:
            st.warning(
                "Snowflake not connected — connect via the sidebar before "
                "running to include the In Table / Match Count scrape-check.",
                icon="❄️",
            )

# ── Auto-restore URL list after page refresh ─────────────────────────────────
if "urls" not in st.session_state and URLLIST_FILE.exists():
    st.session_state.urls = load_urllist()

# ── Checkpoint resume banner ──────────────────────────────────────────────────
resume = False
if any_checkpoint_exists() and URLLIST_FILE.exists():
    all_urls      = load_urllist()
    saved_batches = get_saved_batch_indices()
    total_saved   = sum(len(load_checkpoint(i)) for i in saved_batches)
    num_batches   = (len(all_urls) + BATCH_SIZE - 1) // BATCH_SIZE

    st.warning(
        f"**Interrupted crawl detected** — "
        f"checkpoints found for batch(es): **{saved_batches}** "
        f"({total_saved:,} URLs saved across {len(saved_batches)} batch(es) "
        f"of {num_batches} total)."
    )
    if all_urls:
        progress_val = min(total_saved / len(all_urls), 1.0)
        st.progress(
            progress_val,
            text=f"{progress_val:.0%} complete before interruption",
        )

    col_r, col_f = st.columns(2)
    resume = col_r.button("Resume from checkpoint", use_container_width=True,
                          type="primary")
    if col_f.button("Start fresh", use_container_width=True):
        clear_all_checkpoints()
        st.rerun()

# ── Run ───────────────────────────────────────────────────────────────────────
if "urls" in st.session_state:
    urls        = st.session_state.urls
    total_urls  = len(urls)
    num_batches = (total_urls + BATCH_SIZE - 1) // BATCH_SIZE

    st.divider()
    st.write(f"**{total_urls:,} URLs ready to crawl**")

    if num_batches > 1:
        st.info(
            f"{total_urls:,} URLs split into **{num_batches} batches** of up to "
            f"{BATCH_SIZE:,}. Each batch crawls, checks Snowflake, and produces its "
            f"own Excel file — available for download as soon as that batch finishes."
        )

    st.divider()
    mode = st.radio(
        "**Run mode**",
        options=["Full — Platform URL cleanup + Pass 2 check",
                 "Simple — Check URLs as-is, no cleanup"],
        index=0,
        horizontal=True,
        help=(
            "**Full mode:** strips version segments, derives Platform URLs, "
            "and runs a second-pass fetch to resolve where each platform URL lands.\n\n"
            "**Simple mode:** crawls the URL list exactly as provided — "
            "no version cleaning, no Platform URL derivation, no Pass 2."
        ),
    )
    platform_check = mode.startswith("Full")

    if not platform_check:
        st.info(
            "**Simple mode selected** — URLs will be checked as-is. "
            "Platform URL and Platform Redirect URL columns will be blank in the report.",
            icon="⚡",
        )

    st.divider()

    sf_connected = "sf_conn" in st.session_state
    if not sf_connected:
        st.warning(
            "**Snowflake not connected** — the scrape-check (In Table / Match Count) "
            "will be skipped. Connect via the sidebar before running if you need it.",
            icon="❄️",
        )
        run_btn = st.button(
            "Ready, set, go & RUN! (no Snowflake check)",
            type="primary",
            use_container_width=True,
        )
    else:
        st.success("Snowflake connected — scrape-check will run automatically.")
        run_btn = st.button(
            "Ready, set, go & RUN!",
            type="primary",
            use_container_width=True,
        )

    if run_btn or resume:
        timestamp        = now()
        completed_paths: List[Path] = []

        concepts_table = st.session_state.get("sf_concepts_table", sf_concepts_table)

        url_batches = [
            urls[i : i + BATCH_SIZE]
            for i in range(0, total_urls, BATCH_SIZE)
        ]

        st.divider()
        st.subheader("Downloads")
        placeholders = {}
        for i in range(1, num_batches + 1):
            placeholders[i] = st.empty()

        st.divider()

        for batch_idx, batch_urls in enumerate(url_batches, start=1):
            if resume and not load_checkpoint(batch_idx) and batch_idx in get_saved_batch_indices():
                st.info(f"Batch {batch_idx} already complete — skipping.")
                continue

            path = process_batch(
                batch_urls           = batch_urls,
                batch_idx            = batch_idx,
                num_batches          = num_batches,
                timestamp            = timestamp,
                resume               = resume,
                platform_check       = platform_check,
                download_placeholder = placeholders[batch_idx],
                concepts_table       = concepts_table,
            )
            if path:
                completed_paths.append(path)

        if completed_paths:
            st.session_state.completed_paths = completed_paths

        clear_all_checkpoints()
        st.balloons()
        st.success("All batches complete!")

# ── Persistent download section ───────────────────────────────────────────────
if "completed_paths" in st.session_state:
    st.divider()
    st.subheader("Downloads")
    for p in st.session_state.completed_paths:
        if Path(p).exists():
            with open(p, "rb") as f:
                st.download_button(
                    label=f"Download {Path(p).name}",
                    data=f,
                    file_name=Path(p).name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"dl_persist_{Path(p).name}",
                )
